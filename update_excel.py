#!/usr/bin/env python3
"""
QIPP Excel Export Generator
Patches both Excel templates with roster data and
generates roster_export.js for the HTML dashboard.

Requirements: pip install openpyxl
Usage:        python update_excel.py
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, date, timedelta
import base64, io

# ================================================================
#  ROSTER  — keep in sync with your HTML rosterData.roster array
# ================================================================
ROSTER = [
    {"id":1,"name":"Izhar Ali","fullName":"IZHAR ALI","empId":"500440","position":"60001678","joiningDate":"2025-04-07","nationality":"Pakistani","iqama":"2572283154","employmentType":"3rd Party","company":"Elite","color":"crew-lightviolet","leaves":[{"start":"2026-04-21","end":"2026-05-02","type":"Applied on SAP"}]},
    {"id":2,"name":"Purushothaman","fullName":"PURUSHOTHAMAN DEVARAJ","empId":"500438","position":"60003288","joiningDate":"2025-04-22","nationality":"Indian","iqama":"2563002282","employmentType":"3rd Party","company":"Elite","color":"crew-lightviolet","leaves":[{"start":"2026-04-27","end":"2026-05-07","type":"Applied on SAP"}]},
    {"id":3,"name":"Juma Khan","fullName":"Juma Khan","empId":"2028","position":"60001512","joiningDate":"2013-05-29","nationality":"Pakistani","iqama":"2349465506","employmentType":"Direct Hire","company":"NOMAC","color":"crew-green","leaves":[{"start":"2026-04-29","end":"2026-05-14","type":"Applied on SAP"}]},
    {"id":4,"name":"Mustafa Salem","fullName":"Mustafa Salem Mustafa","empId":"2024","position":"60001655","joiningDate":"2013-05-28","nationality":"Egyptian","iqama":"2349387247","employmentType":"Direct Hire","company":"NOMAC","color":"crew-red","leaves":[{"start":"2026-05-19","end":"2026-06-19","type":"Applied on SAP"}]},
    {"id":5,"name":"Ahmed Alsaqoor","fullName":"AHMAD ALSQOOR","empId":"3973","position":"60001682","joiningDate":"2023-09-03","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightviolet","leaves":[{"start":"2026-05-19","end":"2026-06-11","type":"Applied on SAP"}]},
    {"id":6,"name":"Syed Shahnawaz Ahmed","fullName":"Syed Shahnawaz Ahmed","empId":"2025","position":"60006876","joiningDate":"2013-05-29","nationality":"Pakistani","iqama":"2350058760","employmentType":"Direct Hire","company":"NOMAC","color":"crew-red","leaves":[{"start":"2026-05-20","end":"2026-06-04","type":"Applied on SAP"}]},
    {"id":7,"name":"Abdul Hameed","fullName":"ABDUL HAMEED ABDULRASHEED","empId":"1119","position":"60001507","joiningDate":"2008-05-07","nationality":"Pakistani","iqama":"2258980339","employmentType":"Direct Hire","company":"NOMAC","color":"crew-grey","leaves":[{"start":"2026-05-20","end":"2026-06-10","type":"Applied on SAP"}]},
    {"id":8,"name":"Waleed Fayad","fullName":"Walid Elshahhat Hussein Fayad","empId":"501388","position":"60009705","joiningDate":"2025-07-05","nationality":"Egyptian","iqama":"2610007615","employmentType":"Direct Hire","company":"NOMAC","color":"crew-green","leaves":[{"start":"2026-05-31","end":"2026-06-23","type":"Applied on SAP"}]},
    {"id":9,"name":"Mark Ramirez","fullName":"Mark Anthony Ramirez","empId":"2237","position":"60001527","joiningDate":"2014-02-04","nationality":"Filipino","iqama":"2362010163","employmentType":"3rd Party","company":"Elite","color":"crew-lightviolet","leaves":[{"start":"2026-05-31","end":"2026-06-15","type":"Applied on SAP"}]},
    {"id":10,"name":"Ahmed Fathy","fullName":"Ahmed Fathy Ibrahim AbduelKader","empId":"2726","position":"60004081","joiningDate":"2016-11-02","nationality":"Egyptian","iqama":"2427484262","employmentType":"3rd Party","company":"Elite","color":"crew-lightblue","leaves":[{"start":"2026-06-06","end":"2026-06-14","type":"Applied on SAP"}]},
    {"id":11,"name":"Moustafa Elansari","fullName":"Moustafa Elansary Hewaidy","empId":"2038","position":"60003419","joiningDate":"2013-05-31","nationality":"Egyptian","iqama":"2349203808","employmentType":"Direct Hire","company":"NOMAC","color":"crew-red","leaves":[{"start":"2026-06-16","end":"2026-06-28","type":"Applied on SAP"}]},
    {"id":12,"name":"Veera Prasad","fullName":"Veera Venkata Prasad Vaka","empId":"2294","position":"60001643","joiningDate":"2014-04-16","nationality":"Indian","iqama":"2365111703","employmentType":"Direct Hire","company":"NOMAC","color":"crew-green","leaves":[{"start":"2026-07-06","end":"2026-08-06","type":"Applied on SAP"}]},
    {"id":13,"name":"Lakshmy Prasad","fullName":"LAKSHMI APPALA RAMA DURGA PRASAD","empId":"501234","position":"60001517","joiningDate":"2025-06-29","nationality":"Indian","iqama":"2609943960","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightblue","leaves":[{"start":"2026-07-14","end":"2026-08-10","type":"Applied on SAP"}]},
    {"id":14,"name":"Prathapan","fullName":"SOMANATHAN NAIR PRATHAPAN","empId":"3672","position":"60010950","joiningDate":"2025-05-05","nationality":"Indian","iqama":"2283141766","employmentType":"3rd Party","company":"Haka","color":"crew-lightblue","leaves":[{"start":"2026-07-22","end":"2026-08-18","type":"Applied on SAP"}]},
    {"id":15,"name":"Zaid Almarri","fullName":"Zaid Hadi Almarri","empId":"502667","position":"60001549","joiningDate":"2026-02-03","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-red","leaves":[{"start":"2026-08-01","end":"2026-08-15","type":"Planned"}]},
    {"id":16,"name":"Sami Hamdan","fullName":"Sami Hamdan Dasan Alharbi","empId":"2364","position":"60001972","joiningDate":"2014-08-03","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-yellow","leaves":[{"start":"2026-09-01","end":"2026-09-15","type":"Planned"}]},
    {"id":17,"name":"Fawaz Al Qahtani","fullName":"Fawaz Mari Saeed Alqahtani","empId":"2954","position":"60001526","joiningDate":"2018-10-07","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightblue","leaves":[{"start":"2026-09-16","end":"2026-09-30","type":"Planned"}]},
    {"id":18,"name":"Abdulwahab Al-Shehab","fullName":"Abdulwahab Mohammed Al Shehab","empId":"2202","position":"60003421","joiningDate":"2013-12-16","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightviolet","leaves":[{"start":"2026-10-01","end":"2026-10-15","type":"Planned"}]},
    {"id":19,"name":"Shaheer Yousaf","fullName":"Shaheer Yousaf","empId":"2711","position":"60009706","joiningDate":"2016-09-20","nationality":"Pakistani","iqama":"2421775384","employmentType":"3rd Party","company":"Elite","color":"crew-green","leaves":[{"start":"2026-10-16","end":"2026-10-30","type":"Planned"}]},
    {"id":20,"name":"Mohamed Al Dossary","fullName":"Mohammad AL Dossary","empId":"3865","position":"60001638","joiningDate":"2023-08-20","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightorange","leaves":[{"start":"2026-11-01","end":"2026-11-15","type":"Planned"}]},
    {"id":21,"name":"Abdullah Altulayhi","fullName":"ABDULLAH ALTULAYHI","empId":"501369","position":"60003306","joiningDate":"2025-04-27","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightorange","leaves":[{"start":"2026-11-16","end":"2026-11-30","type":"Planned"}]},
    {"id":22,"name":"Bakr Abdulmajeed","fullName":"Bakr Abdulmajed Alkhabeerani","empId":"3617","position":"60004628","joiningDate":"2025-08-05","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightorange","leaves":[]},
    {"id":23,"name":"Saad Salem Al Hajiri","fullName":"Saad Salem Al Hajri","empId":"4101","position":"60005369","joiningDate":"2026-01-01","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightorange","leaves":[]},
    {"id":24,"name":"Abdulhadi Mohammed Saleh","fullName":"Abdulhadi Mohammed AlMohammed Saleh","empId":"4095","position":"60005376","joiningDate":"2026-01-01","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightorange","leaves":[]},
    {"id":25,"name":"Abdullah Al Amri","fullName":"Abdullah Abdulrahman Alamri","empId":"2363","position":"60001545","joiningDate":"2014-08-03","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-red","leaves":[]},
    {"id":26,"name":"Kanaka Naga Srinivasu Kolli","fullName":"Kanaka Naga Srinivasa Kolli","empId":"500446","position":"60006712","joiningDate":"2013-06-09","nationality":"Indian","iqama":"2349971628","employmentType":"Direct Hire","company":"NOMAC","color":"crew-red","leaves":[]},
    {"id":27,"name":"Ahmed Fawaz","fullName":"AHMED MOUSTAFA AMIN FAWAZ","empId":"","position":"","joiningDate":"2026-02-08","nationality":"Egyptian","iqama":"","employmentType":"Out Source","company":"Jadarah","color":"crew-grey","leaves":[]},
    {"id":28,"name":"Rashed Alhajri","fullName":"Rashed Ghalib Alhajri","empId":"2369","position":"60001508","joiningDate":"2014-08-13","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightblue","leaves":[]},
    {"id":29,"name":"Saravanakumar Madhaiyan","fullName":"Saravanakumar Madhaiyan","empId":"2035","position":"60001520","joiningDate":"2013-05-30","nationality":"Indian","iqama":"2349737847","employmentType":"3rd Party","company":"Elite","color":"crew-lightviolet","leaves":[]},
    {"id":30,"name":"Abdullah Mohiddin Alshahrani","fullName":"Abdullah Mohiddin Alshahrani","empId":"3251","position":"60001516","joiningDate":"2022-11-06","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightblue","leaves":[]},
    {"id":31,"name":"Mohammed Mulhim","fullName":"Mohammed Fahad Al Mulhim","empId":"3262","position":"60001524","joiningDate":"2021-04-18","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightorange","leaves":[]},
    {"id":32,"name":"Mohammed Afnan","fullName":"MOHAMMAD AFNAN","empId":"500439","position":"60001521","joiningDate":"2025-04-07","nationality":"Pakistani","iqama":"2569309392","employmentType":"3rd Party","company":"Elite","color":"crew-lightviolet","leaves":[]},
    {"id":33,"name":"Abdelaziz Al Harbi","fullName":"Abdulaziz Alharbi","empId":"3511","position":"60001525","joiningDate":"2022-03-16","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightorange","leaves":[]},
    {"id":34,"name":"Amjad Alruwaytie","fullName":"Amjad Muflih AlRuwaytie","empId":"4091","position":"60005380","joiningDate":"2026-01-01","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightorange","leaves":[]},
    {"id":35,"name":"Abderahman Al Anezie","fullName":"Abdulrahman Shaiban Al Anazi","empId":"4098","position":"60005372","joiningDate":"2026-01-01","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightorange","leaves":[]},
    {"id":36,"name":"Alaa Alrefae","fullName":"Alaa Abdullah Alrefaei","empId":"502680","position":"60001548","joiningDate":"2026-02-08","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-red","leaves":[]},
    {"id":37,"name":"Abdullah Al Hajri","fullName":"Abdullah Faleh Al hajri","empId":"2665","position":"60001635","joiningDate":"2016-06-16","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightorange","leaves":[]},
    {"id":38,"name":"Yasser Althuiqeb","fullName":"Yaser Althuqaib","empId":"501370","position":"60009704","joiningDate":"2025-05-11","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightblue","leaves":[]},
    {"id":39,"name":"Norbie Cruz","fullName":"Norbie Vianzon Cruz","empId":"2283","position":"60001518","joiningDate":"2014-03-31","nationality":"Filipino","iqama":"2364501045","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightviolet","leaves":[]},
    {"id":40,"name":"Khalid Saleh Khalousi","fullName":"Khaled Khalosy","empId":"501885","position":"60003287","joiningDate":"2025-08-10","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightorange","leaves":[]},
    {"id":41,"name":"Saad Al Enize","fullName":"Seed Alenizi","empId":"3229","position":"60001619","joiningDate":"2021-01-16","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightorange","leaves":[]},
    {"id":42,"name":"Abderahman Al Baqmi","fullName":"Abdulrahman Shabib Al Baqami","empId":"4096","position":"60005375","joiningDate":"2026-01-01","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightorange","leaves":[]},
    {"id":43,"name":"Ali Al Qahtani","fullName":"Ali Mashabab Al Qahtani","empId":"4108","position":"60005382","joiningDate":"2026-01-01","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightorange","leaves":[]},
    {"id":44,"name":"Saleh Mohammed Al Amri","fullName":"Saleh Mohammed Saleh Alamri","empId":"2941","position":"60001563","joiningDate":"2018-08-26","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightblue","leaves":[]},
    {"id":45,"name":"Hassan Arshad","fullName":"Hassan Arshad","empId":"2512","position":"60001509","joiningDate":"2015-09-08","nationality":"Pakistani","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightblue","leaves":[]},
    {"id":46,"name":"Rajesh Muniasamy","fullName":"RAJESH MUNIASAMY","empId":"2492","position":"60001602","joiningDate":"2024-06-03","nationality":"Indian","iqama":"2577251883","employmentType":"3rd Party","company":"Elite","color":"crew-lightviolet","leaves":[]},
    {"id":47,"name":"Mohammed Al Ghamdi","fullName":"Mohammed Al Ghamdi","empId":"2820","position":"60001626","joiningDate":"2017-06-04","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-green","leaves":[]},
    {"id":48,"name":"Bader AlSubait","fullName":"Bader Ibrahim Abdulrahman Alsubeet","empId":"2912","position":"60001681","joiningDate":"2018-04-22","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightorange","leaves":[]},
    {"id":49,"name":"Faris Al Dawasri","fullName":"Faris Shaya Al Dawsari","empId":"4097","position":"60005374","joiningDate":"2026-01-01","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightorange","leaves":[]},
    {"id":50,"name":"Mohamed Al Hakami","fullName":"Mohammed Hassan Hakami","empId":"4081","position":"60005390","joiningDate":"2026-01-01","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightorange","leaves":[]},
    {"id":51,"name":"Saad Al Shahrani","fullName":"Saad Mohammed Al Shahrani","empId":"4094","position":"60005377","joiningDate":"2026-01-01","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightorange","leaves":[]},
    {"id":52,"name":"Mohammad Algarni","fullName":"MOHAMMED ABDULLAH AL GARNI","empId":"3142","position":"60010949","joiningDate":"2025-07-20","nationality":"Saudi Arabian","iqama":"","employmentType":"Direct Hire","company":"NOMAC","color":"crew-lightblue","leaves":[]},
]

# ================================================================
#  COLOR MAP  (CSS class → ARGB for openpyxl)
# ================================================================
COLOR_MAP = {
    'crew-red':         'FFFF7C80',
    'crew-yellow':      'FFFFFF00',
    'crew-green':       'FF92D050',
    'crew-lightviolet': 'FFD9B3FF',
    'crew-lightblue':   'FF9BC2E6',
    'crew-lightorange': 'FFFFCC99',
    'crew-grey':        'FFD9D9D9',
}

# ================================================================
#  COLUMN MAP  — leave date → Excel column number
# ================================================================
#  Jan 2026  : days 1-31  → cols  7-37  (one column per day)
#  Feb 2026  : days 1-28  → cols 38-65  (one column per day)
#  cols 66-80: empty gap
#  Mar-Dec   : milestones 1,6,11,16,21,26 → 6 cols per month
#  Jan 2027  : milestones → cols 141-146
MILESTONE_DAYS = [1, 6, 11, 16, 21, 26]
MONTH_STARTS   = {3:81, 4:87, 5:93, 6:99, 7:105, 8:111,
                  9:117, 10:123, 11:129, 12:135}

LEAVE_COL_MAP: dict[tuple, int] = {}
for d in range(1, 32):
    LEAVE_COL_MAP[(2026, 1, d)] = 6 + d
for d in range(1, 29):
    LEAVE_COL_MAP[(2026, 2, d)] = 37 + d
for m, sc in MONTH_STARTS.items():
    for i, d in enumerate(MILESTONE_DAYS):
        LEAVE_COL_MAP[(2026, m, d)] = sc + i
for i, d in enumerate(MILESTONE_DAYS):
    LEAVE_COL_MAP[(2027, 1, d)] = 141 + i


def get_leave_cols(start_str: str, end_str: str) -> list[int]:
    """Return all Excel column numbers spanned by [start_str, end_str]."""
    start = datetime.strptime(start_str, '%Y-%m-%d').date()
    end   = datetime.strptime(end_str,   '%Y-%m-%d').date()
    cols: set[int] = set()

    # Jan/Feb: daily columns
    cur = max(start, date(2026, 1, 1))
    feb_end = date(2026, 2, 28)
    while cur <= min(end, feb_end):
        key = (cur.year, cur.month, cur.day)
        if key in LEAVE_COL_MAP:
            cols.add(LEAVE_COL_MAP[key])
        cur += timedelta(days=1)

    # Mar 2026 → Jan 2027: milestone columns
    for (y, m, d), col in LEAVE_COL_MAP.items():
        if (y == 2026 and m >= 3) or (y == 2027 and m == 1):
            try:
                check = date(y, m, d)
                if start <= check <= end:
                    cols.add(col)
            except ValueError:
                pass

    return sorted(cols)


def _int_empid(val) -> str:
    """Normalise an Emp ID cell value to a plain string."""
    if val is None:
        return ''
    if isinstance(val, float):
        return str(int(val))
    return str(val).strip()


# ================================================================
#  1. UPDATE MANPOWER FILE
# ================================================================
def update_manpower(template_path: str) -> bytes:
    wb = load_workbook(template_path)
    ws = wb['QIPP']

    # Build fast lookup: empId (str) → person
    by_id: dict[str, dict] = {
        str(p['empId']).strip(): p
        for p in ROSTER if str(p.get('empId', '')).strip()
    }

    for row in range(3, ws.max_row + 1):
        raw_id = ws.cell(row, 6).value          # Column F = Emp ID
        emp_id = _int_empid(raw_id)
        if not emp_id or emp_id not in by_id:
            continue

        p = by_id[emp_id]

        # G = Employee Name
        ws.cell(row, 7).value = p['fullName']

        # H = Position #
        if p.get('position'):
            try:
                ws.cell(row, 8).value = int(p['position'])
            except (ValueError, TypeError):
                ws.cell(row, 8).value = p['position']

        # I = Joining Date
        try:
            ws.cell(row, 9).value = datetime.strptime(p['joiningDate'], '%Y-%m-%d')
        except (ValueError, KeyError):
            pass

        # J = Nationality
        ws.cell(row, 10).value = p['nationality']

        # K = National / Iqama #
        iqama = str(p.get('iqama', '')).strip()
        if iqama:
            try:
                ws.cell(row, 11).value = int(iqama)
            except ValueError:
                ws.cell(row, 11).value = iqama
        else:
            ws.cell(row, 11).value = None

        # L = Employment Type
        ws.cell(row, 12).value = p['employmentType']

        # M = Company
        ws.cell(row, 13).value = p['company']

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ================================================================
#  2. UPDATE LEAVE FILE
# ================================================================
NO_FILL = PatternFill(fill_type=None)

def _normalize(s: str) -> str:
    return str(s).lower().strip() if s else ''


def update_leave(template_path: str) -> bytes:
    wb = load_workbook(template_path)
    ws = wb['Operation ']          # note the trailing space

    # Build name lookup (several name variants per person)
    name_map: dict[str, dict] = {}
    for p in ROSTER:
        for key in [p['fullName'], p['name']]:
            name_map[_normalize(key)] = p

    for row in range(3, ws.max_row + 1):
        cell_name = ws.cell(row, 6).value      # Column F = Names
        if not cell_name:
            continue

        norm = _normalize(cell_name)

        # Match: exact first, then substring
        person = name_map.get(norm)
        if person is None:
            for key, p in name_map.items():
                if key and (key in norm or norm in key):
                    person = p
                    break

        if person is None:
            continue

        # Clear all existing fills in day columns 7-146
        for col in range(7, 147):
            try:
                ws.cell(row, col).fill = NO_FILL
            except Exception:
                pass

        # Apply new leave fills
        rgb  = COLOR_MAP.get(person.get('color', ''), 'FFFFCC99')
        fill = PatternFill(start_color=rgb, end_color=rgb, fill_type='solid')

        for leave in person.get('leaves', []):
            for col in get_leave_cols(leave['start'], leave['end']):
                ws.cell(row, col).fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ================================================================
#  3. MAIN — generate files + roster_export.js
# ================================================================
MANPOWER_TEMPLATE = 'Manpower-QIPP-2026-April-update.xlsx'
LEAVE_TEMPLATE    = 'Copy-of-QIPP-Annual-leave-2026.xlsx'
MANPOWER_OUT      = 'Manpower QIPP 2026 April update.xlsx'
LEAVE_OUT         = 'QIPP Annual Leave 2026.xlsx'
JS_OUT            = 'roster_export.js'


def main():
    print("⏳ Processing Manpower file …")
    manpower_bytes = update_manpower(MANPOWER_TEMPLATE)

    print("⏳ Processing Annual Leave file …")
    leave_bytes = update_leave(LEAVE_TEMPLATE)

    # Save physical Excel files
    with open(MANPOWER_OUT, 'wb') as f:
        f.write(manpower_bytes)
    with open(LEAVE_OUT, 'wb') as f:
        f.write(leave_bytes)

    # Base64-encode for HTML embedding
    manpower_b64 = base64.b64encode(manpower_bytes).decode()
    leave_b64    = base64.b64encode(leave_bytes).decode()

    js_content = (
        "// Auto-generated by update_excel.py — do not edit manually\n"
        f'const MANPOWER_B64 = "{manpower_b64}";\n'
        f'const LEAVE_B64    = "{leave_b64}";\n'
    )
    with open(JS_OUT, 'w', encoding='utf-8') as f:
        f.write(js_content)

    print(f"✅  {MANPOWER_OUT}")
    print(f"✅  {LEAVE_OUT}")
    print(f"✅  {JS_OUT}  ← import this in your HTML")


if __name__ == '__main__':
    main()
